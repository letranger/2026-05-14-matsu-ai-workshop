"""
產出 30 份學生考卷的「正確答案 Excel」——作為 Gemini CLI demo 的對照組。

跑完 generate_exam_pdfs.py 後跑這個，產出 學生考卷-正解.xlsx。
重用同一份題目定義、同一份學生名單、同一個 seed 公式。
"""
from __future__ import annotations
import sys
from pathlib import Path

# 匯入 generate_exam_pdfs 裡的題目／學生／模擬函式
sys.path.insert(0, str(Path(__file__).resolve().parent))
from generate_exam_pdfs import QUESTIONS, STUDENTS, simulate_answers, LETTER  # type: ignore

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "學生考卷-正解.xlsx"


def grade_label(score: int) -> str:
    if score >= 90: return "優"
    if score >= 80: return "甲"
    if score >= 70: return "乙"
    if score >= 60: return "丙"
    return "丁"


# ---------- 計算每位學生的作答與分數 ----------
records = []
for s in STUDENTS:
    picks, correct, score = simulate_answers(seed=s["seat"] * 17 + 31)
    records.append({
        "seat": s["seat"],
        "name": s["name"],
        "picks": picks,            # 每題選的選項 index (0-3)
        "correct_count": correct,
        "score": score,
        "grade": grade_label(score),
    })

# ---------- 建 Excel ----------
wb = Workbook()

# === Sheet 1: 學生成績總表 ===
ws1 = wb.active
ws1.title = "學生成績"

thin = Side(border_style="thin", color="888888")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="305496")
header_font = Font(bold=True, color="FFFFFF", size=11)
center = Alignment(horizontal="center", vertical="center")
right = Alignment(horizontal="right", vertical="center")

# 標題列
headers = ["座號", "姓名"] + [f"Q{i+1}" for i in range(10)] + ["答對題數", "總分", "等第"]
for c, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# 資料列
for r, rec in enumerate(records, 2):
    ws1.cell(row=r, column=1, value=rec["seat"]).alignment = center
    ws1.cell(row=r, column=2, value=rec["name"]).alignment = center
    for q_i, picked in enumerate(rec["picks"]):
        ans_idx = QUESTIONS[q_i]["ans"]
        is_correct = (picked == ans_idx)
        cell_val = LETTER[picked]
        cell = ws1.cell(row=r, column=3 + q_i, value=cell_val)
        cell.alignment = center
        if not is_correct:
            cell.fill = PatternFill("solid", fgColor="FFD6D6")  # 淡紅
            cell.font = Font(color="C00000", bold=True)
    ws1.cell(row=r, column=13, value=rec["correct_count"]).alignment = center
    score_cell = ws1.cell(row=r, column=14, value=rec["score"])
    score_cell.alignment = center
    score_cell.font = Font(bold=True)
    if rec["score"] < 60:
        score_cell.fill = PatternFill("solid", fgColor="FFC7CE")
    elif rec["score"] >= 80:
        score_cell.fill = PatternFill("solid", fgColor="C6EFCE")
    grade_cell = ws1.cell(row=r, column=15, value=rec["grade"])
    grade_cell.alignment = center
    grade_cell.font = Font(bold=True)

    for c in range(1, 16):
        ws1.cell(row=r, column=c).border = border

# 統計列：每題答對率
stats_row = len(records) + 2
label_cell = ws1.cell(row=stats_row, column=1, value="答對率")
label_cell.font = Font(bold=True)
label_cell.fill = PatternFill("solid", fgColor="FFF2CC")
ws1.merge_cells(start_row=stats_row, start_column=1, end_row=stats_row, end_column=2)
label_cell.alignment = center

for q_i in range(10):
    correct_n = sum(1 for r in records if r["picks"][q_i] == QUESTIONS[q_i]["ans"])
    rate = correct_n / len(records)
    cell = ws1.cell(row=stats_row, column=3 + q_i, value=rate)
    cell.number_format = "0.0%"
    cell.alignment = center
    cell.fill = PatternFill("solid", fgColor="FFF2CC")
    if rate < 0.5:
        cell.font = Font(color="C00000", bold=True)
    elif rate >= 0.8:
        cell.font = Font(color="00802B", bold=True)

# 統計列：班級平均
avg = sum(r["score"] for r in records) / len(records)
ws1.cell(row=stats_row, column=13, value=f"平均").font = Font(bold=True)
ws1.cell(row=stats_row, column=13).alignment = center
ws1.cell(row=stats_row, column=13).fill = PatternFill("solid", fgColor="FFF2CC")
avg_cell = ws1.cell(row=stats_row, column=14, value=round(avg, 1))
avg_cell.font = Font(bold=True)
avg_cell.alignment = center
avg_cell.fill = PatternFill("solid", fgColor="FFF2CC")

# 凍結首列
ws1.freeze_panes = "C2"

# 欄寬
widths = [6, 10] + [5] * 10 + [10, 8, 6]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# === Sheet 2: 題目與正解 ===
ws2 = wb.create_sheet("題目與正解")
ws2["A1"] = "題號"
ws2["B1"] = "題目"
ws2["C1"] = "(A)"
ws2["D1"] = "(B)"
ws2["E1"] = "(C)"
ws2["F1"] = "(D)"
ws2["G1"] = "正解"
ws2["H1"] = "答對率"

for c in range(1, 9):
    cell = ws2.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

for q_i, q in enumerate(QUESTIONS):
    r = q_i + 2
    ws2.cell(row=r, column=1, value=q_i + 1).alignment = center
    ws2.cell(row=r, column=2, value=q["q"])
    for opt_i, opt in enumerate(q["options"]):
        cell = ws2.cell(row=r, column=3 + opt_i, value=opt)
        if opt_i == q["ans"]:
            cell.fill = PatternFill("solid", fgColor="C6EFCE")
            cell.font = Font(bold=True, color="00802B")
    ws2.cell(row=r, column=7, value=LETTER[q["ans"]]).alignment = center
    ws2.cell(row=r, column=7).font = Font(bold=True)
    correct_n = sum(1 for rec in records if rec["picks"][q_i] == q["ans"])
    rate_cell = ws2.cell(row=r, column=8, value=correct_n / len(records))
    rate_cell.number_format = "0.0%"
    rate_cell.alignment = center
    if rate_cell.value < 0.5:
        rate_cell.font = Font(color="C00000", bold=True)
    for c in range(1, 9):
        ws2.cell(row=r, column=c).border = border
        if c != 2:
            ws2.cell(row=r, column=c).alignment = center

ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 50
for col in "CDEF":
    ws2.column_dimensions[col].width = 18
ws2.column_dimensions["G"].width = 6
ws2.column_dimensions["H"].width = 9

# === Sheet 3: 班級統計 ===
ws3 = wb.create_sheet("班級統計")
stats = [
    ("學生人數", len(records)),
    ("班級平均", round(sum(r["score"] for r in records) / len(records), 1)),
    ("最高分", max(r["score"] for r in records)),
    ("最低分", min(r["score"] for r in records)),
    ("優（90+）", sum(1 for r in records if r["score"] >= 90)),
    ("甲（80-89）", sum(1 for r in records if 80 <= r["score"] < 90)),
    ("乙（70-79）", sum(1 for r in records if 70 <= r["score"] < 80)),
    ("丙（60-69）", sum(1 for r in records if 60 <= r["score"] < 70)),
    ("丁（< 60）", sum(1 for r in records if r["score"] < 60)),
    ("及格率", f"{sum(1 for r in records if r['score'] >= 60) / len(records) * 100:.1f}%"),
]
for r, (k, v) in enumerate(stats, 1):
    ws3.cell(row=r, column=1, value=k).font = Font(bold=True)
    ws3.cell(row=r, column=2, value=v)
ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 12

wb.save(OUT)
print(f"✓ 已產出 {OUT}")
print(f"   工作表：學生成績 / 題目與正解 / 班級統計")
print(f"   班級平均：{round(sum(r['score'] for r in records) / len(records), 1)}")
